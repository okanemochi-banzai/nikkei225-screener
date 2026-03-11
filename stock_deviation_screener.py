#!/usr/bin/env python3
"""
個別株 理想乖離スクリーナー (Individual Stock Ideal Deviation Screener)
======================================================================
日経225 / ダウ30 / NASDAQ100 の全銘柄に対して
25日移動平均線乖離率と統計的 -2σ / -3σ を算出。
買い場に近い銘柄を一目で把握できるExcel + PNGを出力。

GitHub Actions対応: 毎営業日自動実行 → output/ に結果コミット
"""

import yfinance as yf
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import FancyBboxPatch
from matplotlib.colors import LinearSegmentedColormap
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import time
import os
import json
import warnings
import traceback

warnings.filterwarnings("ignore")

# ─── 設定 ─────────────────────────────────────
SMA_PERIOD = 25
OUTPUT_DIR = os.environ.get("OUTPUT_DIR", "output")

# ─── 銘柄リスト ───────────────────────────────────
# ダウ30
DOW30 = {
    "AAPL": "Apple", "AMGN": "Amgen", "AMZN": "Amazon", "AXP": "American Express",
    "BA": "Boeing", "CAT": "Caterpillar", "CRM": "Salesforce", "CSCO": "Cisco",
    "CVX": "Chevron", "DIS": "Disney", "DOW": "Dow Inc", "GS": "Goldman Sachs",
    "HD": "Home Depot", "HON": "Honeywell", "IBM": "IBM", "JNJ": "Johnson & Johnson",
    "JPM": "JPMorgan Chase", "KO": "Coca-Cola", "MCD": "McDonald's", "MMM": "3M",
    "MRK": "Merck", "MSFT": "Microsoft", "NKE": "Nike", "NVDA": "NVIDIA",
    "PG": "Procter & Gamble", "SHW": "Sherwin-Williams", "TRV": "Travelers",
    "UNH": "UnitedHealth", "V": "Visa", "WMT": "Walmart",
}

# NASDAQ100
NASDAQ100 = {
    "AAPL": "Apple", "ABNB": "Airbnb", "ADBE": "Adobe", "ADI": "Analog Devices",
    "ADP": "ADP", "ADSK": "Autodesk", "AEP": "American Electric Power",
    "AMAT": "Applied Materials", "AMD": "AMD", "AMGN": "Amgen",
    "AMZN": "Amazon", "ANSS": "ANSYS", "APP": "AppLovin", "ARM": "Arm Holdings",
    "ASML": "ASML", "AVGO": "Broadcom", "AZN": "AstraZeneca",
    "BIIB": "Biogen", "BKNG": "Booking Holdings", "BKR": "Baker Hughes",
    "CDNS": "Cadence Design", "CDW": "CDW", "CEG": "Constellation Energy",
    "CHTR": "Charter Comm", "CMCSA": "Comcast", "COIN": "Coinbase",
    "COST": "Costco", "CPRT": "Copart", "CRWD": "CrowdStrike",
    "CSCO": "Cisco", "CSGP": "CoStar Group", "CTAS": "Cintas",
    "CTSH": "Cognizant", "DASH": "DoorDash", "DDOG": "Datadog",
    "DLTR": "Dollar Tree", "DXCM": "DexCom",
    "EA": "Electronic Arts", "EXC": "Exelon",
    "FANG": "Diamondback Energy", "FAST": "Fastenal", "FTNT": "Fortinet",
    "GEHC": "GE HealthCare", "GFS": "GlobalFoundries", "GILD": "Gilead",
    "GOOG": "Alphabet C", "GOOGL": "Alphabet A",
    "HON": "Honeywell", "IDXX": "IDEXX Labs", "ILMN": "Illumina",
    "INTC": "Intel", "INTU": "Intuit", "ISRG": "Intuitive Surgical",
    "KDP": "Keurig Dr Pepper", "KHC": "Kraft Heinz", "KLAC": "KLA Corp",
    "LIN": "Linde", "LRCX": "Lam Research", "LULU": "Lululemon",
    "MAR": "Marriott", "MCHP": "Microchip Tech", "MDB": "MongoDB",
    "MDLZ": "Mondelez", "MELI": "MercadoLibre", "META": "Meta Platforms",
    "MNST": "Monster Beverage", "MRVL": "Marvell Tech", "MSFT": "Microsoft",
    "MU": "Micron", "NFLX": "Netflix", "NVDA": "NVIDIA",
    "NXPI": "NXP Semi", "ODFL": "Old Dominion Freight",
    "ON": "ON Semiconductor", "ORLY": "O'Reilly Auto", "PANW": "Palo Alto Networks",
    "PAYX": "Paychex", "PCAR": "PACCAR", "PDD": "PDD Holdings",
    "PEP": "PepsiCo", "PLTR": "Palantir", "PYPL": "PayPal",
    "QCOM": "Qualcomm", "REGN": "Regeneron", "ROP": "Roper Technologies",
    "ROST": "Ross Stores", "SBUX": "Starbucks", "SMCI": "Super Micro Computer",
    "SNPS": "Synopsys", "TEAM": "Atlassian", "TMUS": "T-Mobile",
    "TSLA": "Tesla", "TTD": "The Trade Desk", "TTWO": "Take-Two",
    "TXN": "Texas Instruments", "VRSK": "Verisk Analytics", "VRTX": "Vertex Pharma",
    "WBD": "Warner Bros Discovery", "WDAY": "Workday", "XEL": "Xcel Energy",
    "ZS": "Zscaler",
}

# 日経225 (証券コード.T)
NIKKEI225 = {
    "1332.T": "日本水産", "1333.T": "マルハニチロ",
    "1605.T": "INPEX", "1721.T": "コムシスHD",
    "1801.T": "大成建設", "1802.T": "大林組", "1803.T": "清水建設",
    "1808.T": "長谷工", "1812.T": "鹿島建設", "1925.T": "大和ハウス",
    "1928.T": "積水ハウス", "1963.T": "日揮HD",
    "2002.T": "日清製粉G", "2121.T": "MIXI",
    "2175.T": "SMS", "2267.T": "ヤクルト",
    "2413.T": "エムスリー", "2432.T": "ディー・エヌ・エー",
    "2501.T": "サッポロHD", "2502.T": "アサヒGHD",
    "2503.T": "キリンHD", "2531.T": "宝HD",
    "2768.T": "双日", "2801.T": "キッコーマン",
    "2802.T": "味の素", "2871.T": "ニチレイ",
    "2914.T": "JT", "3003.T": "ヒューリック",
    "3049.T": "エノテカ(注:代替)", "3086.T": "Jフロント",
    "3088.T": "マツキヨココカラ", "3092.T": "ZOZO",
    "3099.T": "三越伊勢丹HD", "3105.T": "日清紡HD",
    "3116.T": "トヨタ紡織", "3231.T": "野村不動産HD",
    "3244.T": "サムティ(注:代替)", "3289.T": "東急不動産HD",
    "3349.T": "コスモス薬品", "3382.T": "セブン&アイ",
    "3401.T": "帝人", "3402.T": "東レ",
    "3405.T": "クラレ", "3407.T": "旭化成",
    "3436.T": "SUMCO", "3659.T": "ネクソン",
    "3861.T": "王子HD", "3863.T": "日本製紙",
    "4004.T": "レゾナックHD", "4005.T": "住友化学",
    "4021.T": "日産化学", "4042.T": "東ソー",
    "4043.T": "トクヤマ", "4061.T": "デンカ",
    "4063.T": "信越化学", "4151.T": "協和キリン",
    "4183.T": "三井化学", "4188.T": "三菱ケミカルG",
    "4208.T": "UBE", "4324.T": "電通G",
    "4385.T": "メルカリ", "4452.T": "花王",
    "4502.T": "武田薬品", "4503.T": "アステラス製薬",
    "4506.T": "住友ファーマ", "4507.T": "塩野義製薬",
    "4519.T": "中外製薬", "4523.T": "エーザイ",
    "4528.T": "小野薬品", "4543.T": "テルモ",
    "4568.T": "第一三共", "4578.T": "大塚HD",
    "4612.T": "日本ペイントHD", "4631.T": "DIC",
    "4661.T": "OLC", "4684.T": "オービック(注:代替)",
    "4689.T": "LINEヤフー", "4704.T": "トレンドマイクロ",
    "4716.T": "日本オラクル(注:代替)", "4751.T": "サイバーエージェント",
    "4755.T": "楽天グループ", "4901.T": "富士フイルムHD",
    "4902.T": "コニカミノルタ", "4911.T": "資生堂",
    "5019.T": "出光興産", "5020.T": "ENEOS HD",
    "5101.T": "横浜ゴム", "5108.T": "ブリヂストン",
    "5201.T": "AGC", "5214.T": "日本電気硝子",
    "5233.T": "太平洋セメント", "5301.T": "東海カーボン",
    "5332.T": "TOTO", "5333.T": "日本ガイシ",
    "5334.T": "日本特殊陶業", "5401.T": "日本製鉄",
    "5406.T": "神戸製鋼所", "5411.T": "JFE HD",
    "5413.T": "日新製鋼(注:代替)", "5541.T": "大平洋金属(注:代替)",
    "5631.T": "日本製鋼所", "5703.T": "日本軽金属HD",
    "5706.T": "三井金属鉱業", "5707.T": "東邦亜鉛",
    "5711.T": "三菱マテリアル", "5713.T": "住友金属鉱山",
    "5714.T": "DOWAホールディングス", "5801.T": "古河電気工業",
    "5802.T": "住友電気工業", "5803.T": "フジクラ",
    "5831.T": "しずおかFG", "6098.T": "リクルートHD",
    "6103.T": "オークマ", "6113.T": "アマダ",
    "6141.T": "DMG森精機", "6146.T": "ディスコ",
    "6178.T": "日本郵政", "6273.T": "SMC",
    "6301.T": "小松製作所", "6302.T": "住友重機械",
    "6305.T": "日立建機", "6326.T": "クボタ",
    "6361.T": "荏原製作所", "6367.T": "ダイキン工業",
    "6370.T": "栗田工業", "6471.T": "日本精工",
    "6472.T": "NTN", "6473.T": "ジェイテクト",
    "6479.T": "ミネベアミツミ", "6501.T": "日立製作所",
    "6503.T": "三菱電機", "6504.T": "富士電機",
    "6506.T": "安川電機", "6594.T": "ニデック",
    "6645.T": "オムロン", "6674.T": "GSユアサ",
    "6701.T": "NEC", "6702.T": "富士通",
    "6723.T": "ルネサスエレクトロニクス", "6724.T": "セイコーエプソン",
    "6752.T": "パナソニックHD", "6753.T": "シャープ",
    "6758.T": "ソニーG", "6762.T": "TDK",
    "6770.T": "アルプスアルパイン", "6841.T": "横河電機",
    "6857.T": "アドバンテスト", "6861.T": "キーエンス",
    "6902.T": "デンソー", "6920.T": "レーザーテック",
    "6952.T": "カシオ計算機", "6954.T": "ファナック",
    "6971.T": "京セラ", "6976.T": "太陽誘電",
    "6981.T": "村田製作所", "7003.T": "三井E&S",
    "7004.T": "日立造船(注:代替)", "7011.T": "三菱重工業",
    "7012.T": "川崎重工業", "7013.T": "IHI",
    "7186.T": "コンコルディアFG", "7201.T": "日産自動車",
    "7202.T": "いすゞ自動車", "7203.T": "トヨタ自動車",
    "7205.T": "日野自動車(注:代替)", "7211.T": "三菱自動車",
    "7261.T": "マツダ", "7267.T": "ホンダ",
    "7269.T": "スズキ", "7270.T": "SUBARU",
    "7272.T": "ヤマハ発動機", "7309.T": "シマノ",
    "7733.T": "オリンパス", "7735.T": "SCREENホールディングス",
    "7741.T": "HOYA", "7751.T": "キヤノン",
    "7752.T": "リコー", "7762.T": "シチズン時計",
    "7832.T": "バンダイナムコHD", "7911.T": "凸版印刷(注:代替)",
    "7912.T": "大日本印刷", "7951.T": "ヤマハ",
    "7974.T": "任天堂", "8001.T": "伊藤忠商事",
    "8002.T": "丸紅", "8015.T": "豊田通商",
    "8028.T": "ファミリーマート(注:代替)", "8031.T": "三井物産",
    "8035.T": "東京エレクトロン", "8053.T": "住友商事",
    "8058.T": "三菱商事", "8233.T": "高島屋",
    "8252.T": "丸井グループ", "8253.T": "クレディセゾン",
    "8267.T": "イオン", "8303.T": "新生銀行(注:代替)",
    "8304.T": "あおぞら銀行(注:代替)", "8306.T": "三菱UFJ FG",
    "8308.T": "りそなHD", "8309.T": "三井住友トラストHD",
    "8316.T": "三井住友FG", "8331.T": "千葉銀行",
    "8354.T": "ふくおかFG", "8355.T": "静岡銀行(注:代替)",
    "8411.T": "みずほFG", "8591.T": "オリックス",
    "8601.T": "大和証券G", "8604.T": "野村HD",
    "8630.T": "SOMPOホールディングス", "8697.T": "日本取引所G",
    "8725.T": "MS&ADインシュアランスG", "8750.T": "第一生命HD",
    "8766.T": "東京海上HD", "8795.T": "T&D HD",
    "8801.T": "三井不動産", "8802.T": "三菱地所",
    "8804.T": "東京建物", "8830.T": "住友不動産",
    "9001.T": "東武鉄道", "9005.T": "東急",
    "9007.T": "小田急電鉄", "9008.T": "京王電鉄",
    "9009.T": "京成電鉄", "9020.T": "東日本旅客鉄道",
    "9021.T": "西日本旅客鉄道", "9022.T": "東海旅客鉄道",
    "9024.T": "西武HD(注:代替)", "9064.T": "ヤマトHD",
    "9101.T": "日本郵船", "9104.T": "商船三井",
    "9107.T": "川崎汽船", "9201.T": "日本航空",
    "9202.T": "ANA HD", "9301.T": "三菱倉庫",
    "9432.T": "NTT", "9433.T": "KDDI",
    "9434.T": "ソフトバンク", "9501.T": "東京電力HD",
    "9502.T": "中部電力", "9503.T": "関西電力",
    "9531.T": "東京ガス", "9532.T": "大阪ガス",
    "9602.T": "東宝", "9613.T": "NTTデータG",
    "9735.T": "セコム", "9766.T": "コナミG",
    "9783.T": "ベネッセHD(注:代替)", "9843.T": "ニトリHD",
    "9983.T": "ファーストリテイリング", "9984.T": "ソフトバンクG",
}


def get_zone(dev, s2l, s3l, s2u, s3u):
    if dev <= s3l: return "STRONG BUY"
    elif dev <= s2l: return "BUY ZONE"
    elif dev <= 0: return "MILD DIP"
    elif dev <= s2u: return "NEUTRAL"
    elif dev <= s3u: return "OVERBOUGHT"
    else: return "EXTREME HIGH"


def calc_deviation(ticker_str, name, market):
    """1銘柄の理想乖離を計算"""
    try:
        ticker = yf.Ticker(ticker_str)
        df = ticker.history(period="max")
        if df.empty or len(df) < SMA_PERIOD + 100:
            return None

        # 配当情報取得（yfinanceのdividendYieldは日本株で異常値が多いため、
        # dividendRate ÷ 株価 で自前計算する）
        div_consec_years = 0
        try:
            info = ticker.info
            div_per_share = info.get("dividendRate") or 0.0

            # dividendRateがない場合、直近1年の配当履歴から計算
            divs = pd.Series(dtype=float)
            try:
                divs = ticker.dividends
                if len(divs) > 0 and not div_per_share:
                    one_year_ago = divs.index[-1] - pd.Timedelta(days=400)
                    recent_divs = divs[divs.index >= one_year_ago]
                    if len(recent_divs) > 0:
                        div_per_share = float(recent_divs.sum())
            except:
                pass

            # 連続増配年数の計算（年間配当を年ごとに合算し、前年より増えている連続年数）
            if len(divs) > 0:
                try:
                    annual_divs = divs.groupby(divs.index.year).sum()
                    if len(annual_divs) >= 2:
                        # 直近の不完全年は除外する可能性があるため、最新年を除外
                        current_year = datetime.now().year
                        if annual_divs.index[-1] == current_year and len(annual_divs) >= 3:
                            annual_divs = annual_divs.iloc[:-1]
                        # 逆順にたどって連続増配をカウント
                        years_counted = 0
                        for i in range(len(annual_divs) - 1, 0, -1):
                            if annual_divs.iloc[i] > annual_divs.iloc[i-1]:
                                years_counted += 1
                            else:
                                break
                        div_consec_years = years_counted
                except:
                    pass

            div_yield = 0.0  # 株価取得後に計算
        except:
            div_yield = 0.0
            div_per_share = 0.0

        df = df[["Close"]].copy()
        df.columns = ["close"]
        df["sma25"] = df["close"].rolling(SMA_PERIOD).mean()
        df["deviation"] = ((df["close"] - df["sma25"]) / df["sma25"]) * 100
        df.dropna(inplace=True)

        if len(df) < 200:
            return None

        dev = df["deviation"]
        mean_d = dev.mean()
        std_d = dev.std()
        s2l = mean_d - 2 * std_d
        s3l = mean_d - 3 * std_d
        s2u = mean_d + 2 * std_d
        s3u = mean_d + 3 * std_d

        latest = df.iloc[-1]
        cur_price = latest["close"]
        cur_sma = latest["sma25"]
        cur_dev = latest["deviation"]
        price_2s = cur_sma * (1 + s2l / 100)
        price_3s = cur_sma * (1 + s3l / 100)
        dist_2s = ((cur_price / price_2s) - 1) * 100

        # 配当利回りを自前計算（dividendRate ÷ 現在株価）
        if div_per_share and cur_price > 0:
            div_yield = round(div_per_share / cur_price * 100, 2)
        else:
            div_yield = 0.0

        # -2σ到達時の想定配当利回り
        if div_per_share and price_2s > 0:
            div_at_2s = round(div_per_share / price_2s * 100, 2)
        else:
            div_at_2s = 0.0

        # ========== 追加シグナル ==========
        # 日足全データ（乖離率計算前のdfは加工済みなので、元データを使う）
        df_full = ticker.history(period="max")[["Close", "Open", "High", "Low"]].copy()
        df_full.columns = ["close", "open", "high", "low"]

        # ① 高配当フラグ（日本株4%超、米国株6%超）
        div_threshold = 4.0 if market == "Nikkei225" else 6.0
        flag_high_div = div_yield >= div_threshold

        # ② 過去最高値から半値以下
        ath = df_full["close"].max()
        half_from_ath = cur_price <= ath * 0.5
        drop_from_ath = round((1 - cur_price / ath) * 100, 1) if ath > 0 else 0

        # ③ 25日移動平均線を上抜け（直近で下→上にクロス）
        sma25_prev = df.iloc[-2]["sma25"] if len(df) >= 2 else cur_sma
        close_prev = df.iloc[-2]["close"] if len(df) >= 2 else cur_price
        cross_above_25ma = (close_prev < sma25_prev) and (cur_price >= cur_sma)

        # ④ 月足で久々の陽線（直近月が陽線 かつ その前2ヶ月が陰線）
        try:
            df_monthly = df_full.resample("ME").agg({"open": "first", "close": "last"}).dropna()
            if len(df_monthly) >= 3:
                latest_m = df_monthly.iloc[-1]
                prev1_m = df_monthly.iloc[-2]
                prev2_m = df_monthly.iloc[-3]
                monthly_bullish = (
                    latest_m["close"] > latest_m["open"] and  # 直近月が陽線
                    prev1_m["close"] < prev1_m["open"] and    # 前月が陰線
                    prev2_m["close"] < prev2_m["open"]        # 前々月も陰線
                )
            else:
                monthly_bullish = False
        except:
            monthly_bullish = False

        # ⑤ 手法12: 月足ゴールデンクロス（月足の短期MA(9)が中期MA(24)を上抜け）
        try:
            df_m_close = df_full["close"].resample("ME").last().dropna()
            if len(df_m_close) >= 25:
                m_sma9 = df_m_close.rolling(9).mean()
                m_sma24 = df_m_close.rolling(24).mean()
                # 直近で9MAが24MAを下から上にクロス
                gc_now = m_sma9.iloc[-1] >= m_sma24.iloc[-1]
                gc_prev = m_sma9.iloc[-2] < m_sma24.iloc[-2]
                monthly_golden_cross = gc_now and gc_prev
                # 底打ち判定: 過去6ヶ月の安値から反発中
                m_low_6 = df_m_close.iloc[-6:].min()
                rebounding = cur_price > m_low_6 * 1.05
                method12 = monthly_golden_cross or (gc_now and rebounding and drop_from_ath >= 30)
            else:
                monthly_golden_cross = False
                method12 = False
        except:
            monthly_golden_cross = False
            method12 = False

        # ⑦ セクター情報取得
        SECTOR_JP = {
            "Technology": "テクノロジー", "Financial Services": "金融",
            "Healthcare": "ヘルスケア", "Consumer Cyclical": "一般消費財",
            "Consumer Defensive": "生活必需品", "Energy": "エネルギー",
            "Industrials": "資本財", "Basic Materials": "素材",
            "Communication Services": "通信", "Utilities": "公益",
            "Real Estate": "不動産",
        }
        try:
            sector_en = info.get("sector") or ""
            sector = SECTOR_JP.get(sector_en, sector_en)
            industry = info.get("industry") or ""
        except:
            sector = ""
            industry = ""

        # シグナルまとめ
        signals = []
        if flag_high_div:
            signals.append("高配当")
        if half_from_ath:
            signals.append("半値")
        if cross_above_25ma:
            signals.append("日足25MA上抜け")
        if monthly_bullish:
            signals.append("月足陽転")
        if method12:
            signals.append("月足GC")

        n_signals = len(signals)
        signal_text = "／".join(signals) if signals else "—"

        # 配当利回りキリ番の到達株価（年間配当 ÷ 目標利回り）
        yield_targets = {}
        for pct in [3, 4, 5, 6]:
            if div_per_share > 0:
                yield_targets[f"price_at_{pct}pct"] = round(div_per_share / (pct / 100), 1)
            else:
                yield_targets[f"price_at_{pct}pct"] = 0.0

        zone = get_zone(cur_dev, s2l, s3l, s2u, s3u)

        # ========== 買いスコア (0〜100) ==========
        score = 0

        # (1) -2σへの近さ (最大40点)
        if dist_2s <= 0:
            score += 40
        elif dist_2s <= 3:
            score += 30
        elif dist_2s <= 5:
            score += 20
        elif dist_2s <= 10:
            score += 10

        # (2) -3σ到達ボーナス (5点)
        if cur_dev <= s3l:
            score += 5

        # (3) 配当利回り (最大20点)
        if flag_high_div:
            score += 20
        elif div_yield >= div_threshold * 0.75:
            score += 12
        elif div_yield >= div_threshold * 0.5:
            score += 5

        # (4) テクニカルシグナル (各5点, 最大15点)
        if cross_above_25ma:
            score += 5
        if monthly_bullish:
            score += 5
        if method12:
            score += 5

        # (5) 底値圏ボーナス (最大10点)
        if half_from_ath:
            score += 10
        elif drop_from_ath >= 30:
            score += 5

        # (6) 統計充実度 + 安定性 (最大10点)
        if std_d <= 4 and len(dev) >= 3000:
            score += 10
        elif std_d <= 6 and len(dev) >= 2000:
            score += 5

        score = min(score, 100)

        return {
            "ticker": ticker_str,
            "name": name,
            "market": market,
            "price": round(cur_price, 2),
            "sma25": round(cur_sma, 2),
            "deviation": round(cur_dev, 3),
            "sigma2_lower": round(s2l, 3),
            "sigma3_lower": round(s3l, 3),
            "sigma2_upper": round(s2u, 3),
            "sigma3_upper": round(s3u, 3),
            "price_at_2s": round(price_2s, 1),
            "price_at_3s": round(price_3s, 1),
            "dist_to_2s": round(dist_2s, 2),
            "zone": zone,
            "stat_days": len(dev),
            "std": round(std_d, 3),
            "div_yield": div_yield,
            "div_per_share": round(div_per_share, 2) if div_per_share else 0.0,
            "div_at_2s": div_at_2s,
            "div_consec_years": div_consec_years,
            "buy_score": score,
            "ath": round(ath, 1),
            "drop_from_ath": drop_from_ath,
            "half_from_ath": half_from_ath,
            "cross_above_25ma": cross_above_25ma,
            "monthly_bullish": monthly_bullish,
            "method12": method12,
            "flag_high_div": flag_high_div,
            "sector": sector,
            "industry": industry,
            "signals": signal_text,
            "n_signals": n_signals,
            **yield_targets,
        }
    except Exception as e:
        return None


# ─── メイン処理 ─────────────────────────────────
def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    today = datetime.now().strftime("%Y-%m-%d")

    print("=" * 70)
    print(f"  Individual Stock Ideal Deviation Screener — {today}")
    print("=" * 70)

    # 全銘柄リスト作成
    all_stocks = []
    for t, n in NIKKEI225.items():
        all_stocks.append((t, n, "Nikkei225"))
    for t, n in DOW30.items():
        all_stocks.append((t, n, "Dow30"))
    for t, n in NASDAQ100.items():
        if t not in DOW30:  # 重複排除
            all_stocks.append((t, n, "NASDAQ100"))

    total = len(all_stocks)
    print(f"\n  Total stocks to process: {total}")
    print(f"  Nikkei225: {len(NIKKEI225)} / Dow30: {len(DOW30)} / NASDAQ100(unique): {total - len(NIKKEI225) - len(DOW30)}\n")

    results = []
    errors = []
    for i, (ticker, name, market) in enumerate(all_stocks):
        if (i + 1) % 20 == 0 or i == 0:
            print(f"  Processing {i+1}/{total}... ({ticker})")
        data = calc_deviation(ticker, name, market)
        if data:
            results.append(data)
        else:
            errors.append(ticker)
        # yfinance rate limit対策
        if (i + 1) % 50 == 0:
            time.sleep(2)

    print(f"\n  Completed: {len(results)} success / {len(errors)} failed")
    if errors:
        print(f"  Failed tickers: {', '.join(errors[:20])}{'...' if len(errors)>20 else ''}")

    if not results:
        print("  [ERROR] No results. Exiting.")
        return

    # DataFrameに変換・ソート
    df = pd.DataFrame(results)

    # ⑦ セクター内で唯一下落している銘柄を判定
    df["sector_sole_dip"] = False
    for sector in df["sector"].unique():
        if not sector:
            continue
        sec_df = df[df["sector"] == sector]
        if len(sec_df) < 3:
            continue
        # セクター内で乖離率マイナスが1銘柄だけ
        dipping = sec_df[sec_df["deviation"] < 0]
        others = sec_df[sec_df["deviation"] >= 0]
        if len(dipping) == 1 and len(others) >= 2:
            idx = dipping.index[0]
            df.loc[idx, "sector_sole_dip"] = True
            # シグナルに追加
            old_sig = df.loc[idx, "signals"]
            df.loc[idx, "signals"] = (old_sig + "／セクター唯一安" if old_sig != "—" else "セクター唯一安")
            df.loc[idx, "n_signals"] = df.loc[idx, "n_signals"] + 1

    df.sort_values("dist_to_2s", ascending=True, inplace=True)
    df.reset_index(drop=True, inplace=True)

    # ========== Excel出力 ==========
    print("\n  Generating Excel...")
    wb = Workbook()

    # --- シート1: 全銘柄一覧 (dist_to_2sでソート) ---
    ws_all = wb.active
    ws_all.title = "All Stocks"
    headers = ["順位", "買いスコア", "シグナル", "ゾーン", "コード", "銘柄名", "セクター", "市場",
               "株価", "25日MA", "乖離率%", "-2σまで%",
               "配当利回り%", "配当@-2σ%", "連続増配", "最高値からの下落%",
               "標準偏差", "統計日数"]
    ws_all.append(headers)

    # ヘッダースタイル
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    for col_idx, h in enumerate(headers, 1):
        cell = ws_all.cell(row=1, column=col_idx)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ゾーン色マップ
    zone_fills = {
        "STRONG BUY": PatternFill("solid", fgColor="FF4444"),
        "BUY ZONE": PatternFill("solid", fgColor="FF8C00"),
        "MILD DIP": PatternFill("solid", fgColor="FFF176"),
        "NEUTRAL": PatternFill("solid", fgColor="C8E6C9"),
        "OVERBOUGHT": PatternFill("solid", fgColor="CE93D8"),
        "EXTREME HIGH": PatternFill("solid", fgColor="880E4F"),
    }
    zone_fonts = {
        "STRONG BUY": Font(bold=True, color="FFFFFF", name="Arial"),
        "BUY ZONE": Font(bold=True, color="FFFFFF", name="Arial"),
        "MILD DIP": Font(name="Arial"),
        "NEUTRAL": Font(name="Arial"),
        "OVERBOUGHT": Font(color="FFFFFF", name="Arial"),
        "EXTREME HIGH": Font(bold=True, color="FFFFFF", name="Arial"),
    }

    def write_stock_row(ws, r, idx, row, zone_fills, zone_fonts):
        """1銘柄ぶんのExcel行を書き込む共通関数"""
        c = 1
        ws.cell(row=r, column=c, value=idx); c += 1

        # 買いスコア
        score_cell = ws.cell(row=r, column=c, value=row.get("buy_score", 0))
        sc = row.get("buy_score", 0)
        if sc >= 70:
            score_cell.font = Font(bold=True, color="FF0000", name="Arial")
            score_cell.fill = PatternFill("solid", fgColor="FFEBEE")
        elif sc >= 50:
            score_cell.font = Font(bold=True, color="FF8C00", name="Arial")
            score_cell.fill = PatternFill("solid", fgColor="FFF3E0")
        elif sc >= 30:
            score_cell.font = Font(color="2E7D32", name="Arial")
        c += 1

        # シグナル
        sig_cell = ws.cell(row=r, column=c, value=row.get("signals", "—"))
        n_sig = row.get("n_signals", 0)
        if n_sig >= 3:
            sig_cell.font = Font(bold=True, color="FF0000", name="Arial")
        elif n_sig >= 2:
            sig_cell.font = Font(bold=True, color="FF8C00", name="Arial")
        elif n_sig >= 1:
            sig_cell.font = Font(color="1565C0", name="Arial")
        c += 1

        # ゾーン
        zone_cell = ws.cell(row=r, column=c, value=row["zone"])
        z = row["zone"]
        if z in zone_fills:
            zone_cell.fill = zone_fills[z]
            zone_cell.font = zone_fonts.get(z, Font(name="Arial"))
        c += 1

        ws.cell(row=r, column=c, value=row["ticker"]); c += 1
        ws.cell(row=r, column=c, value=row["name"]); c += 1
        ws.cell(row=r, column=c, value=row.get("sector", "")); c += 1
        ws.cell(row=r, column=c, value=row["market"]); c += 1
        ws.cell(row=r, column=c, value=row["price"]).number_format = '#,##0.00'; c += 1
        ws.cell(row=r, column=c, value=row["sma25"]).number_format = '#,##0.00'; c += 1
        ws.cell(row=r, column=c, value=row["deviation"]).number_format = '0.00"%"'; c += 1

        # -2σまで%
        dist_cell = ws.cell(row=r, column=c, value=row["dist_to_2s"])
        dist_cell.number_format = '0.00"%"'
        if row["dist_to_2s"] <= 0:
            dist_cell.font = Font(bold=True, color="FF0000", name="Arial")
        elif row["dist_to_2s"] <= 3:
            dist_cell.font = Font(bold=True, color="FF8C00", name="Arial")
        c += 1

        # 配当利回り
        div_threshold = 4.0 if row["market"] == "Nikkei225" else 6.0
        div_cell = ws.cell(row=r, column=c, value=row["div_yield"])
        div_cell.number_format = '0.00"%"'
        if row["div_yield"] >= div_threshold:
            div_cell.font = Font(bold=True, color="008000", name="Arial")
        c += 1

        # 配当@-2σ
        div2s_cell = ws.cell(row=r, column=c, value=row["div_at_2s"])
        div2s_cell.number_format = '0.00"%"'
        if row["div_at_2s"] >= div_threshold:
            div2s_cell.font = Font(bold=True, color="008000", name="Arial")
        c += 1

        # 連続増配年数
        consec = row.get("div_consec_years", 0) or 0
        consec_cell = ws.cell(row=r, column=c, value=consec if consec > 0 else "")
        if consec >= 10:
            consec_cell.font = Font(bold=True, color="008000", name="Arial")
            consec_cell.fill = PatternFill("solid", fgColor="E8F5E9")
        elif consec >= 5:
            consec_cell.font = Font(color="2E7D32", name="Arial")
        c += 1

        # 最高値からの下落%
        drop_cell = ws.cell(row=r, column=c, value=row.get("drop_from_ath", 0))
        drop_cell.number_format = '0.0"%"'
        if row.get("drop_from_ath", 0) >= 50:
            drop_cell.font = Font(bold=True, color="FF0000", name="Arial")
        elif row.get("drop_from_ath", 0) >= 30:
            drop_cell.font = Font(color="FF8C00", name="Arial")
        c += 1

        ws.cell(row=r, column=c, value=row["std"]).number_format = '0.000'; c += 1
        ws.cell(row=r, column=c, value=row["stat_days"]).number_format = '#,##0'

    for idx, row in df.iterrows():
        write_stock_row(ws_all, idx + 2, idx + 1, row, zone_fills, zone_fonts)

    # カラム幅調整
    col_widths = [5, 9, 28, 12, 10, 20, 16, 11, 11, 11, 9, 12, 9, 10, 9, 12, 8, 9]
    for i, w in enumerate(col_widths, 1):
        ws_all.column_dimensions[get_column_letter(i)].width = w

    # フリーズペイン
    ws_all.freeze_panes = "A2"
    # オートフィルタ
    ws_all.auto_filter.ref = f"A1:R{len(df)+1}"

    # --- シート2-4: マーケット別 ---
    for market_name in ["Nikkei225", "Dow30", "NASDAQ100"]:
        ws = wb.create_sheet(title=market_name)
        mdf = df[df["market"] == market_name].reset_index(drop=True)
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")

        for idx2, row in mdf.iterrows():
            write_stock_row(ws, idx2 + 2, idx2 + 1, row, zone_fills, zone_fonts)

        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:R{len(mdf)+1}"

    # --- シート5: BUY候補 (dist_to_2s <= 3%) ---
    ws_buy = wb.create_sheet(title="BUY Candidates")
    buy_df = df[df["dist_to_2s"] <= 3.0].reset_index(drop=True)
    ws_buy.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws_buy.cell(row=1, column=col_idx)
        cell.fill = PatternFill("solid", fgColor="B71C1C")
        cell.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        cell.alignment = Alignment(horizontal="center")

    for idx3, row in buy_df.iterrows():
        write_stock_row(ws_buy, idx3 + 2, idx3 + 1, row, zone_fills, zone_fonts)

    for i, w in enumerate(col_widths, 1):
        ws_buy.column_dimensions[get_column_letter(i)].width = w
    ws_buy.freeze_panes = "A2"

    # --- シート6: シグナル点灯銘柄 (1つ以上のシグナルあり) ---
    ws_sig = wb.create_sheet(title="Signals")
    sig_df = df[df["n_signals"] >= 1].sort_values("n_signals", ascending=False).reset_index(drop=True)
    ws_sig.append(headers)
    sig_fill = PatternFill("solid", fgColor="1A237E")
    for col_idx in range(1, len(headers) + 1):
        cell = ws_sig.cell(row=1, column=col_idx)
        cell.fill = sig_fill
        cell.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        cell.alignment = Alignment(horizontal="center")
    for idx4, row in sig_df.iterrows():
        write_stock_row(ws_sig, idx4 + 2, idx4 + 1, row, zone_fills, zone_fonts)
    for i, w in enumerate(col_widths, 1):
        ws_sig.column_dimensions[get_column_letter(i)].width = w
    ws_sig.freeze_panes = "A2"

    xlsx_path = os.path.join(OUTPUT_DIR, f"stock_deviation_{today}.xlsx")
    wb.save(xlsx_path)
    print(f"  Saved: {xlsx_path}")

    # ========== CSV出力（Webダッシュボード用） ==========
    csv_path = os.path.join(OUTPUT_DIR, f"stock_deviation_{today}.csv")
    df.to_csv(csv_path, index=False)
    print(f"  Saved: {csv_path}")

    # ========== PNG: ゾーン分布サマリー ==========
    print("  Generating PNG summary...")
    fig, axes = plt.subplots(1, 3, figsize=(18, 7))

    zone_order = ["STRONG BUY", "BUY ZONE", "MILD DIP", "NEUTRAL", "OVERBOUGHT", "EXTREME HIGH"]
    zone_colors_plt = ["#d32f2f", "#ff8c00", "#fdd835", "#66bb6a", "#ab47bc", "#880e4f"]

    for ax_idx, (market, title) in enumerate([
        ("Nikkei225", "Nikkei 225"), ("Dow30", "Dow 30"), ("NASDAQ100", "NASDAQ 100")
    ]):
        ax = axes[ax_idx]
        mdf = df[df["market"] == market]
        zone_counts = mdf["zone"].value_counts()
        counts = [zone_counts.get(z, 0) for z in zone_order]
        bars = ax.barh(zone_order[::-1], counts[::-1],
                       color=zone_colors_plt[::-1], edgecolor="white", height=0.6)
        for bar, c in zip(bars, counts[::-1]):
            if c > 0:
                ax.text(bar.get_width() + 0.5, bar.get_y() + bar.get_height()/2,
                        str(c), va="center", fontsize=10, fontweight="bold")
        ax.set_title(f"{title}  ({len(mdf)} stocks)", fontsize=13, fontweight="bold", pad=10)
        ax.set_xlim(0, max(counts) * 1.3 + 2)
        ax.tick_params(axis="y", labelsize=9)
        ax.tick_params(axis="x", labelsize=8)
        ax.grid(axis="x", alpha=0.2)

    fig.suptitle(f"Ideal Deviation Zone Distribution — {today}",
                 fontsize=15, fontweight="bold", y=1.02)
    fig.tight_layout()
    png1 = os.path.join(OUTPUT_DIR, "zone_distribution.png")
    fig.savefig(png1, dpi=150, bbox_inches="tight", facecolor="white")
    print(f"  Saved: {png1}")

    # ========== PNG: TOP20 買い候補 ==========
    fig2, ax2 = plt.subplots(figsize=(14, 8))
    top20 = df.head(20).copy()
    top20 = top20.iloc[::-1]  # 逆順で下から表示

    colors = []
    for _, r in top20.iterrows():
        if r["dist_to_2s"] <= 0:
            colors.append("#d32f2f")
        elif r["dist_to_2s"] <= 3:
            colors.append("#ff8c00")
        elif r["dist_to_2s"] <= 5:
            colors.append("#fdd835")
        else:
            colors.append("#90caf9")

    labels = [f"{r['ticker']}  {r['name']}" for _, r in top20.iterrows()]
    bars = ax2.barh(range(len(top20)), top20["dist_to_2s"].values, color=colors,
                    edgecolor="white", height=0.7)

    ax2.axvline(x=0, color="#d32f2f", linewidth=2, linestyle="-", alpha=0.7, label="-2σ line")
    ax2.axvline(x=3, color="#ff8c00", linewidth=1.5, linestyle="--", alpha=0.5, label="3% buffer")

    for bar, val, (_, r) in zip(bars, top20["dist_to_2s"].values, top20.iterrows()):
        label_text = f" {val:+.1f}%  (Dev:{r['deviation']:+.1f}%)"
        ax2.text(bar.get_width() + 0.3, bar.get_y() + bar.get_height()/2,
                 label_text, va="center", fontsize=8, fontweight="bold")

    ax2.set_yticks(range(len(top20)))
    ax2.set_yticklabels(labels, fontsize=9)
    ax2.set_xlabel("Distance to -2σ Buy Zone (%)", fontsize=10)
    ax2.set_title(f"Top 20 Closest to Buy Zone — {today}", fontsize=14, fontweight="bold")
    ax2.legend(fontsize=9, loc="lower right")
    ax2.grid(axis="x", alpha=0.2)
    fig2.tight_layout()
    png2 = os.path.join(OUTPUT_DIR, "top20_buy_candidates.png")
    fig2.savefig(png2, dpi=150, bbox_inches="tight", facecolor="white")
    print(f"  Saved: {png2}")

    plt.close("all")

    # ========== サマリー ==========
    print("\n" + "=" * 70)
    print(f"  ZONE SUMMARY ({today})")
    print("=" * 70)
    for market in ["Nikkei225", "Dow30", "NASDAQ100"]:
        mdf = df[df["market"] == market]
        print(f"\n  [{market}]")
        for z in zone_order:
            cnt = len(mdf[mdf["zone"] == z])
            if cnt > 0:
                print(f"    {z:<16} {cnt:>3} stocks")

    buy_count = len(df[df["zone"].isin(["STRONG BUY", "BUY ZONE"])])
    near_buy = len(df[df["dist_to_2s"] <= 3.0])
    print(f"\n  Total BUY/STRONG BUY: {buy_count}")
    print(f"  Within 3% of -2σ: {near_buy}")
    print("=" * 70)
    print("  Done.\n")


if __name__ == "__main__":
    main()
