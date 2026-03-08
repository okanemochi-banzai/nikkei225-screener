"""
日経225 月足MA乖離率スクリーナー
- 月足5MA/20MA乖離率の算出
- GCシグナル検出（発生・接近）
- 25MA乖離率の5パーセンタイル算出（暴落逆張り基準）
"""

import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
import time
import os

warnings.filterwarnings("ignore")

# =============================================================
# 日経225構成銘柄（定期入替あり・適宜更新のこと）
# =============================================================
NIKKEI225 = {
    "1332": "ニッスイ", "1333": "マルハニチロ",
    "1605": "INPEX", "1721": "コムシスHD",
    "1801": "大成建設", "1802": "大林組",
    "1803": "清水建設", "1808": "長谷工",
    "1812": "鹿島建設", "1925": "大和ハウス",
    "1928": "積水ハウス", "1963": "日揮HD",
    "2002": "日清製粉G", "2269": "明治HD",
    "2282": "日本ハム", "2413": "エムスリー",
    "2432": "ディー・エヌ・エー", "2501": "サッポロHD",
    "2502": "アサヒGHD", "2503": "キリンHD",
    "2531": "宝HD", "2768": "双日",
    "2801": "キッコーマン", "2802": "味の素",
    "2871": "ニチレイ", "2914": "JT",
    "3003": "ヒューリック", "3086": "Jフロント",
    "3088": "マツキヨココカラ", "3092": "ZOZO",
    "3099": "三越伊勢丹HD", "3105": "日清紡HD",
    "3116": "トヨタ紡織", "3231": "野村不動産HD",
    "3283": "日本プロロジス", "3289": "東急不動産HD",
    "3382": "セブン＆アイ", "3401": "帝人",
    "3402": "東レ", "3405": "クラレ",
    "3407": "旭化成", "3436": "SUMCO",
    "3659": "ネクソン", "3861": "王子HD",
    "3863": "日本製紙", "4004": "レゾナック",
    "4005": "住友化学", "4021": "日産化学",
    "4042": "東ソー", "4043": "トクヤマ",
    "4061": "デンカ", "4063": "信越化学",
    "4091": "太陽日酸", "4151": "協和キリン",
    "4183": "三井化学", "4188": "三菱ケミカルG",
    "4204": "積水化学", "4208": "UBE",
    "4307": "野村総合研究所", "4324": "電通G",
    "4385": "メルカリ", "4452": "花王",
    "4502": "武田薬品", "4503": "アステラス製薬",
    "4506": "住友ファーマ", "4507": "塩野義製薬",
    "4519": "中外製薬", "4523": "エーザイ",
    "4528": "小野薬品", "4543": "テルモ",
    "4568": "第一三共", "4578": "大塚HD",
    "4612": "日本ペイントHD", "4631": "DIC",
    "4661": "オリエンタルランド", "4680": "ラウンドワン",
    "4689": "LINEヤフー", "4704": "トレンドマイクロ",
    "4751": "サイバーエージェント", "4755": "楽天G",
    "4901": "富士フイルム", "4902": "コニカミノルタ",
    "4911": "資生堂", "5019": "出光興産",
    "5020": "ENEOS HD", "5101": "横浜ゴム",
    "5108": "ブリヂストン", "5201": "AGC",
    "5202": "板硝子", "5214": "日本電気硝子",
    "5233": "太平洋セメント", "5301": "東海カーボン",
    "5332": "TOTO", "5333": "日本ガイシ",
    "5401": "日本製鉄", "5406": "神戸製鋼",
    "5411": "JFE HD", "5413": "日新製鋼",
    "5541": "大平洋金属", "5631": "日本製鋼所",
    "5703": "日本軽金属HD", "5706": "三井金属",
    "5707": "東邦亜鉛", "5711": "三菱マテリアル",
    "5713": "住友金属鉱山", "5714": "DOWAホールディングス",
    "5801": "古河電工", "5802": "住友電工",
    "5803": "フジクラ", "5901": "東洋製罐GHD",
    "6098": "リクルートHD", "6103": "オークマ",
    "6113": "アマダ", "6141": "DMG森精機",
    "6146": "ディスコ", "6178": "日本郵政",
    "6273": "SMC", "6301": "小松製作所",
    "6302": "住友重機械", "6305": "日立建機",
    "6326": "クボタ", "6361": "荏原製作所",
    "6367": "ダイキン工業", "6370": "栗田工業",
    "6383": "ダイフク", "6471": "日本精工",
    "6472": "NTN", "6473": "ジェイテクト",
    "6479": "ミネベアミツミ", "6501": "日立製作所",
    "6503": "三菱電機", "6504": "富士電機",
    "6506": "安川電機", "6526": "ソシオネクスト",
    "6532": "ベイカレント", "6586": "マキタ",
    "6594": "ニデック", "6617": "東光高岳",
    "6645": "オムロン", "6674": "GSユアサ",
    "6701": "NEC", "6702": "富士通",
    "6723": "ルネサスエレクトロニクス", "6724": "セイコーエプソン",
    "6752": "パナソニックHD", "6753": "シャープ",
    "6758": "ソニーG", "6762": "TDK",
    "6770": "アルプスアルパイン", "6841": "横河電機",
    "6857": "アドバンテスト", "6861": "キーエンス",
    "6869": "シスメックス", "6902": "デンソー",
    "6920": "レーザーテック", "6952": "カシオ",
    "6954": "ファナック", "6976": "太陽誘電",
    "6981": "村田製作所", "6988": "日東電工",
    "7003": "三井E&S", "7004": "日立造船",
    "7011": "三菱重工", "7012": "川崎重工",
    "7013": "IHI", "7014": "名村造船所",
    "7186": "コンコルディアFG", "7201": "日産自動車",
    "7202": "いすゞ", "7203": "トヨタ自動車",
    "7205": "日野自動車", "7211": "三菱自動車",
    "7261": "マツダ", "7267": "ホンダ",
    "7269": "スズキ", "7270": "SUBARU",
    "7272": "ヤマハ発動機", "7309": "シマノ",
    "7412": "アフラック", "7453": "良品計画",
    "7532": "パンパシHD", "7733": "オリンパス",
    "7735": "SCREENホールディングス", "7741": "HOYA",
    "7751": "キヤノン", "7752": "リコー",
    "7762": "シチズン時計", "7832": "バンダイナムコHD",
    "7911": "凸版HD", "7912": "大日本印刷",
    "7951": "ヤマハ", "7974": "任天堂",
    "8001": "伊藤忠", "8002": "丸紅",
    "8015": "豊田通商", "8031": "三井物産",
    "8035": "東京エレクトロン", "8053": "住友商事",
    "8058": "三菱商事", "8113": "ユニ・チャーム",
    "8233": "高島屋", "8252": "丸井G",
    "8253": "クレディセゾン", "8267": "イオン",
    "8303": "SBI新生銀行", "8304": "あおぞら銀行",
    "8306": "三菱UFJ FG", "8308": "りそなHD",
    "8309": "三井住友トラストHD", "8316": "三井住友FG",
    "8331": "千葉銀行", "8354": "ふくおかFG",
    "8355": "静岡銀行", "8411": "みずほFG",
    "8591": "オリックス", "8601": "大和証券G",
    "8604": "野村HD", "8628": "松井証券",
    "8630": "SOMPOホールディングス", "8725": "MS&ADインシュアランスG",
    "8750": "第一生命HD", "8766": "東京海上HD",
    "8795": "T&D HD", "8801": "三井不動産",
    "8802": "三菱地所", "8804": "東京建物",
    "8830": "住友不動産", "9001": "東武鉄道",
    "9005": "東急", "9007": "小田急電鉄",
    "9008": "京王電鉄", "9009": "京成電鉄",
    "9020": "JR東日本", "9021": "JR西日本",
    "9022": "JR東海", "9024": "西武HD",
    "9042": "阪急阪神HD", "9064": "ヤマトHD",
    "9101": "日本郵船", "9104": "商船三井",
    "9107": "川崎汽船", "9202": "ANA HD",
    "9301": "三菱倉庫", "9412": "スカパーJSAT",
    "9432": "NTT", "9433": "KDDI",
    "9434": "ソフトバンク", "9501": "東京電力HD",
    "9502": "中部電力", "9503": "関西電力",
    "9531": "東京ガス", "9532": "大阪ガス",
    "9602": "東宝", "9613": "NTTデータG",
    "9735": "セコム", "9766": "コナミG",
    "9983": "ファーストリテイリング", "9984": "ソフトバンクG",
}


def fetch_stock_data(code: str, name: str) -> dict | None:
    """1銘柄の月足データ取得 & MA指標算出"""
    ticker = f"{code}.T"
    try:
        data = yf.download(ticker, period="5y", interval="1mo", progress=False)
        if len(data) < 20:
            return None

        close = data["Close"].squeeze()
        ma5 = close.rolling(5).mean()
        ma20 = close.rolling(20).mean()
        ma25 = close.rolling(25).mean()

        latest_close = float(close.iloc[-1])
        latest_ma5 = float(ma5.iloc[-1])
        latest_ma20 = float(ma20.iloc[-1])
        divergence = (latest_ma5 - latest_ma20) / latest_ma20 if latest_ma20 != 0 else None

        # GCシグナル判定
        gc_signal = ""
        if len(ma5) >= 2 and len(ma20) >= 2:
            prev_diff = float(ma5.iloc[-2] - ma20.iloc[-2])
            curr_diff = float(latest_ma5 - latest_ma20)
            if not np.isnan(prev_diff) and not np.isnan(curr_diff):
                if prev_diff < 0 and curr_diff >= 0:
                    gc_signal = "★ GC発生"
                elif prev_diff < 0 and latest_ma20 != 0 and abs(curr_diff / latest_ma20) < 0.03:
                    gc_signal = "● GC接近"

        # 25MA乖離率
        latest_ma25 = float(ma25.iloc[-1]) if len(ma25.dropna()) > 0 else None
        div_25ma = (latest_close - latest_ma25) / latest_ma25 if latest_ma25 and latest_ma25 != 0 else None

        # 5パーセンタイル
        pct5 = None
        if len(ma25.dropna()) > 0:
            div_series = (close - ma25) / ma25
            pct5 = float(div_series.dropna().quantile(0.05))

        return {
            "銘柄コード": code,
            "銘柄名": name,
            "直近終値": round(latest_close, 1),
            "月足5MA": round(latest_ma5, 1),
            "月足20MA": round(latest_ma20, 1),
            "MA乖離率(5-20)": round(divergence, 4) if divergence else None,
            "GCシグナル": gc_signal,
            "25MA乖離率": round(div_25ma, 4) if div_25ma else None,
            "25MA乖離5%tile": round(pct5, 4) if pct5 else None,
        }
    except Exception:
        return None


def build_excel(df: pd.DataFrame, output_path: str):
    """DataFrameからフォーマット済みExcelを生成"""
    wb = Workbook()

    # スタイル定義
    NAVY = "1B2A4A"
    BLUE = "2E5090"
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor=BLUE)
    title_font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
    title_fill = PatternFill("solid", fgColor=NAVY)
    data_font = Font(name="Arial", size=10, color="333333")
    gc_font = Font(name="Arial", size=10, color="D32F2F", bold=True)
    near_gc_font = Font(name="Arial", size=10, color="FF9800", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_al = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    alt_fill = PatternFill("solid", fgColor="F5F5F5")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    gc_fill = PatternFill("solid", fgColor="FFEBEE")
    near_fill = PatternFill("solid", fgColor="FFF3E0")

    headers = [
        "銘柄コード", "銘柄名", "直近終値\n(円)", "月足5MA",
        "月足20MA", "MA乖離率\n(5MA-20MA)", "GCシグナル",
        "25MA乖離率\n(対終値)", "25MA乖離\n5%tile",
    ]
    widths = [12, 22, 12, 12, 12, 14, 14, 14, 14]

    def write_header(ws, headers, fill, row=2):
        ws.row_dimensions[row].height = 38
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = hdr_font
            cell.fill = fill
            cell.alignment = center
            cell.border = border

    def write_data(ws, data_df, start_row=3):
        for idx, row_data in data_df.iterrows():
            r = idx + start_row
            is_gc = row_data.get("GCシグナル") == "★ GC発生"
            is_near = row_data.get("GCシグナル") == "● GC接近"
            vals = [row_data.get(h.split("\n")[0].replace("(円)", "").replace("(5MA-20MA)", "(5-20)").replace("(対終値)", "").strip(), "")
                    for h in headers]
            # fallback: use column keys directly
            vals = [
                row_data["銘柄コード"], row_data["銘柄名"], row_data["直近終値"],
                row_data["月足5MA"], row_data["月足20MA"], row_data["MA乖離率(5-20)"],
                row_data.get("GCシグナル", ""), row_data.get("25MA乖離率", None),
                row_data.get("25MA乖離5%tile", None),
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = border
                if is_gc:
                    cell.fill = gc_fill
                    cell.font = gc_font if c == 7 else data_font
                elif is_near:
                    cell.fill = near_fill
                    cell.font = near_gc_font if c == 7 else data_font
                else:
                    cell.fill = alt_fill if idx % 2 == 0 else white_fill
                    cell.font = data_font
                cell.alignment = center if c != 2 else left_al
                if c in [6, 8, 9]:
                    cell.number_format = "0.00%"
                elif c in [3, 4, 5]:
                    cell.number_format = "#,##0.0"

    # --- Sheet 1: 全銘柄一覧 ---
    ws1 = wb.active
    ws1.title = "MA乖離率スクリーニング"
    ws1.sheet_properties.tabColor = BLUE

    now_str = datetime.now().strftime("%Y/%m/%d")
    ws1.merge_cells("A1:I1")
    ws1.cell(row=1, column=1, value=f"日経225 月足MA乖離率スクリーニング（{now_str}）").font = title_font
    ws1.cell(row=1, column=1).fill = title_fill
    ws1.cell(row=1, column=1).alignment = center
    ws1.row_dimensions[1].height = 34
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    sorted_df = df.sort_values("MA乖離率(5-20)", ascending=True).reset_index(drop=True)
    write_header(ws1, headers, hdr_fill)
    write_data(ws1, sorted_df)
    ws1.freeze_panes = "A3"
    ws1.auto_filter.ref = f"A2:I{len(sorted_df)+2}"

    # --- Sheet 2: GC候補 ---
    ws2 = wb.create_sheet("①GC候補銘柄")
    ws2.sheet_properties.tabColor = "E53935"
    gc_df = sorted_df[sorted_df["GCシグナル"].isin(["★ GC発生", "● GC接近"])].reset_index(drop=True)

    ws2.merge_cells("A1:I1")
    ws2.cell(row=1, column=1, value=f"①戦略: GC発生・接近銘柄（{len(gc_df)}銘柄）").font = title_font
    ws2.cell(row=1, column=1).fill = PatternFill("solid", fgColor="C62828")
    ws2.cell(row=1, column=1).alignment = center
    ws2.row_dimensions[1].height = 34
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    write_header(ws2, headers, PatternFill("solid", fgColor="C62828"))
    write_data(ws2, gc_df)
    ws2.freeze_panes = "A3"

    # --- Sheet 3: 暴落逆張り基準 ---
    ws3 = wb.create_sheet("②暴落逆張り基準値")
    ws3.sheet_properties.tabColor = "FF9800"
    crash_df = df.sort_values("25MA乖離率", ascending=True).reset_index(drop=True)

    crash_headers = [
        "銘柄コード", "銘柄名", "直近終値\n(円)",
        "25MA乖離率\n(現在)", "25MA乖離\n5%tile",
        "打診買い目安\n(5%tile)", "本格買い目安\n(×1.5)",
        "月足5MA", "月足20MA",
    ]
    crash_widths = [12, 22, 12, 14, 14, 14, 14, 12, 12]

    ws3.merge_cells("A1:I1")
    ws3.cell(row=1, column=1, value="②戦略: 25MA乖離率 & 5%tile基準値一覧").font = title_font
    ws3.cell(row=1, column=1).fill = PatternFill("solid", fgColor="E65100")
    ws3.cell(row=1, column=1).alignment = center
    ws3.row_dimensions[1].height = 34
    for i, w in enumerate(crash_widths, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    write_header(ws3, crash_headers, PatternFill("solid", fgColor="E65100"))

    for idx, row_data in crash_df.iterrows():
        r = idx + 3
        pct5 = row_data["25MA乖離5%tile"]
        vals = [
            row_data["銘柄コード"], row_data["銘柄名"], row_data["直近終値"],
            row_data["25MA乖離率"], pct5,
            pct5,
            round(pct5 * 1.5, 4) if pct5 else None,
            row_data["月足5MA"], row_data["月足20MA"],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row=r, column=c, value=v)
            cell.border = border
            cell.font = data_font
            cell.fill = alt_fill if idx % 2 == 0 else white_fill
            cell.alignment = center if c != 2 else left_al
            if c in [4, 5, 6, 7]:
                cell.number_format = "0.00%"
            elif c in [3, 8, 9]:
                cell.number_format = "#,##0.0"

    ws3.freeze_panes = "A3"
    ws3.auto_filter.ref = f"A2:I{len(crash_df)+2}"

    wb.save(output_path)


def main():
    print(f"📊 日経225 MA乖離率スクリーナー 実行開始 ({datetime.now().strftime('%Y-%m-%d %H:%M')})")
    print(f"   対象: {len(NIKKEI225)}銘柄\n")

    results = []
    errors = []

    for i, (code, name) in enumerate(NIKKEI225.items(), 1):
        result = fetch_stock_data(code, name)
        if result:
            results.append(result)
        else:
            errors.append(f"{code} {name}")

        if i % 50 == 0:
            print(f"  ... {i}/{len(NIKKEI225)} 完了")
            time.sleep(2)

    print(f"\n✅ データ取得完了: 成功={len(results)}, エラー={len(errors)}")

    if errors:
        print(f"⚠️  取得失敗: {', '.join(errors[:10])}")

    df = pd.DataFrame(results)

    # output ディレクトリ作成
    os.makedirs("output", exist_ok=True)
    filename = f"output/nikkei225_MA_{datetime.now().strftime('%Y%m%d')}.xlsx"

    build_excel(df, filename)
    print(f"📁 Excel出力: {filename}")

    # サマリー表示
    gc_count = len(df[df["GCシグナル"] == "★ GC発生"])
    near_count = len(df[df["GCシグナル"] == "● GC接近"])
    print(f"\n--- サマリー ---")
    print(f"  GC発生: {gc_count}銘柄")
    print(f"  GC接近: {near_count}銘柄")
    print(f"  25MA乖離率 最低: {df['25MA乖離率'].min():.2%}")
    print(f"🎉 完了!")


if __name__ == "__main__":
    main()
